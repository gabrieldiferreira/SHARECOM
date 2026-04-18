from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone
import hashlib
import io

router = APIRouter(prefix="/export", tags=["export"])


class TransactionExport(BaseModel):
    merchant_name: str
    total_amount: float
    category: str
    transaction_type: str
    payment_method: str
    transaction_date: str
    destination_institution: Optional[str] = None
    note: Optional[str] = None


class ExportRequest(BaseModel):
    transactions: List[TransactionExport]
    report_title: str = "Relatório de Transações"
    report_type: str = "general"  # general, category, monthly, payment_method
    customer_name: Optional[str] = None
    customer_cpf_cnpj: Optional[str] = None


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
@router.post("/excel")
def export_excel(req: ExportRequest):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---------- Main sheet ----------
    ws = wb.active
    ws.title = "Transações"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = req.report_title
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Customer name
    if req.customer_name:
        ws.merge_cells("A2:H2")
        name_cell = ws["A2"]
        name_cell.value = f"Cliente: {req.customer_name}"
        name_cell.font = Font(name="Calibri", bold=True, size=11, color="374151")
        name_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 22

    # Subtitle
    sub_row = 3 if req.customer_name else 2
    ws.merge_cells(f"A{sub_row}:H{sub_row}")
    sub_cell = ws[f"A{sub_row}"]
    sub_cell.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub_cell.font = Font(name="Calibri", size=10, color="6B7280")
    sub_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[sub_row].height = 20

    header_row = 5 if req.customer_name else 4
    headers = ["Data", "Comerciante", "Categoria", "Tipo", "Método", "Instituição", "Valor (R$)", "Nota"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    inflow_font = Font(name="Calibri", color="059669")
    outflow_font = Font(name="Calibri", color="DC2626")
    money_fmt = '#,##0.00'

    data_start = header_row + 1
    for i, tx in enumerate(req.transactions):
        row = i + data_start
        try:
            dt = datetime.fromisoformat(tx.transaction_date.replace("Z", "+00:00"))
            date_str = dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            date_str = tx.transaction_date

        values = [
            date_str,
            tx.merchant_name,
            tx.category,
            "Entrada" if tx.transaction_type == "Inflow" else "Saída",
            tx.payment_method,
            tx.destination_institution or "",
            tx.total_amount,
            tx.note or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col_idx == 7:
                cell.number_format = money_fmt
                cell.font = inflow_font if tx.transaction_type == "Inflow" else outflow_font

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = len(headers[col_idx - 1])
        for row in range(data_start, len(req.transactions) + data_start):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    # ---------- Summary sheet ----------
    ws2 = wb.create_sheet("Resumo")
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Resumo Financeiro"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="1E3A5F")
    ws2["A1"].alignment = Alignment(horizontal="center")

    inflow_total = sum(t.total_amount for t in req.transactions if t.transaction_type == "Inflow")
    outflow_total = sum(t.total_amount for t in req.transactions if t.transaction_type == "Outflow")

    summary_headers = ["Métrica", "Valor"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    summary_data = [
        ("Total de Transações", len(req.transactions)),
        ("Total Entradas (R$)", inflow_total),
        ("Total Saídas (R$)", outflow_total),
        ("Saldo (R$)", inflow_total - outflow_total),
    ]
    for i, (label, val) in enumerate(summary_data):
        ws2.cell(row=4 + i, column=1, value=label).font = Font(name="Calibri", bold=True)
        c = ws2.cell(row=4 + i, column=2, value=val)
        if isinstance(val, float):
            c.number_format = money_fmt

    # Category breakdown
    cat_totals: dict[str, float] = {}
    for tx in req.transactions:
        cat_totals[tx.category] = cat_totals.get(tx.category, 0) + tx.total_amount

    ws2.cell(row=10, column=1, value="Categoria").font = header_font
    ws2.cell(row=10, column=1).fill = header_fill
    ws2.cell(row=10, column=2, value="Total (R$)").font = header_font
    ws2.cell(row=10, column=2).fill = header_fill
    for i, (cat, total) in enumerate(sorted(cat_totals.items(), key=lambda x: -x[1])):
        ws2.cell(row=11 + i, column=1, value=cat)
        c = ws2.cell(row=11 + i, column=2, value=total)
        c.number_format = money_fmt

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Helpers – Brazilian formatting
# ---------------------------------------------------------------------------
def fmt_real(valor: float) -> str:
    """Formata valor em R$ padrão brasileiro (Receita Federal)."""
    if valor < 0:
        return f"(R$ {abs(valor):,.2f})".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_cpf_cnpj(raw: Optional[str]) -> str:
    """Formata CPF (11 dígitos) ou CNPJ (14 dígitos) com pontuação."""
    if not raw:
        return ""
    digits = "".join(c for c in raw if c.isdigit())
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    if len(digits) == 14:
        return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"
    return raw  # already formatted or unknown length


def _parse_tx_date(raw: str) -> datetime:
    """Best-effort parse of transaction date string."""
    try:
        # Parse ISO format, handling 'Z' suffix
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        # Ensure it's naive for comparison with datetime.now() and other potentially naive dates
        if dt.tzinfo is not None:
            # Convert to UTC first to preserve the absolute time, then make it naive
            return dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt
    except Exception:
        return datetime.now()


def _generate_hash(cliente: str, data_geracao: str, total_transacoes: int) -> str:
    payload = f"{cliente}{data_geracao}{total_transacoes}"
    return hashlib.md5(payload.encode()).hexdigest()[:16].upper()


# ---------------------------------------------------------------------------
# PDF – Padrão Receita Federal
# ---------------------------------------------------------------------------
@router.post("/pdf")
def export_pdf(req: ExportRequest):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    # -- colour palette (spec) ------------------------------------------------
    CLR_HEADER_BG   = colors.HexColor("#F3F4F6")
    CLR_ROW_ALT     = colors.HexColor("#FAFAFA")
    CLR_SALDO_BG    = colors.HexColor("#FEF3C7")
    CLR_TEXT_PRI     = colors.HexColor("#1F2937")
    CLR_TEXT_SEC     = colors.HexColor("#4B5563")
    CLR_BORDER       = colors.HexColor("#E5E7EB")

    # -- pre-compute values ---------------------------------------------------
    now = datetime.now()
    data_geracao = now.strftime("%d/%m/%Y %H:%M")
    hash_doc = _generate_hash(
        req.customer_name or "", data_geracao, len(req.transactions),
    )

    sorted_tx = sorted(req.transactions, key=lambda t: _parse_tx_date(t.transaction_date))

    inflow_total  = sum(t.total_amount for t in req.transactions if t.transaction_type == "Inflow")
    outflow_total = sum(t.total_amount for t in req.transactions if t.transaction_type == "Outflow")
    saldo = inflow_total - outflow_total

    # Determine report period from transactions
    if sorted_tx:
        dt_first = _parse_tx_date(sorted_tx[0].transaction_date)
        dt_last  = _parse_tx_date(sorted_tx[-1].transaction_date)
        periodo = f"{dt_first.strftime('%d/%m/%Y')} a {dt_last.strftime('%d/%m/%Y')}"
    else:
        periodo = data_geracao.split()[0]

    # -- styles ---------------------------------------------------------------
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "RFTitle", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=16, textColor=CLR_TEXT_PRI, alignment=TA_CENTER, spaceAfter=2,
    )
    header_info_style = ParagraphStyle(
        "RFHeaderInfo", parent=styles["Normal"], fontName="Helvetica",
        fontSize=9, textColor=CLR_TEXT_PRI, alignment=TA_LEFT, spaceAfter=1,
        leading=12,
    )
    section_style = ParagraphStyle(
        "RFSection", parent=styles["Heading2"], fontName="Helvetica-Bold",
        fontSize=12, textColor=CLR_TEXT_PRI, spaceBefore=8 * mm, spaceAfter=3 * mm,
    )
    cell_style = ParagraphStyle(
        "RFCell", parent=styles["Normal"], fontName="Helvetica",
        fontSize=9, leading=11, textColor=CLR_TEXT_PRI,
    )
    cell_center = ParagraphStyle("RFCellC", parent=cell_style, alignment=TA_CENTER)
    cell_right  = ParagraphStyle("RFCellR", parent=cell_style, alignment=TA_RIGHT)
    cell_left   = ParagraphStyle("RFCellL", parent=cell_style, alignment=TA_LEFT)
    cell_bold   = ParagraphStyle(
        "RFCellB", parent=cell_style, fontName="Helvetica-Bold",
    )
    cell_bold_right = ParagraphStyle(
        "RFCellBR", parent=cell_style, fontName="Helvetica-Bold", alignment=TA_RIGHT,
    )
    obs_style = ParagraphStyle(
        "RFObs", parent=styles["Normal"], fontName="Helvetica",
        fontSize=9, textColor=CLR_TEXT_SEC, leading=13, spaceAfter=6,
    )

    # -- common table style builder -------------------------------------------
    def _base_table_style(num_rows: int, extra=None):
        s = [
            ("GRID",            (0, 0), (-1, -1), 0.5, CLR_BORDER),
            ("BACKGROUND",      (0, 0), (-1, 0),  CLR_HEADER_BG),
            ("FONTNAME",        (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",        (0, 0), (-1, 0),  10),
            ("FONTSIZE",        (0, 1), (-1, -1), 9),
            ("VALIGN",          (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",      (0, 0), (-1, -1), 3 * mm),
            ("BOTTOMPADDING",   (0, 0), (-1, -1), 3 * mm),
            ("LEFTPADDING",     (0, 0), (-1, -1), 2 * mm),
            ("RIGHTPADDING",    (0, 0), (-1, -1), 2 * mm),
            ("TEXTCOLOR",       (0, 0), (-1, -1), CLR_TEXT_PRI),
            ("ROWBACKGROUNDS",  (0, 1), (-1, -1), [colors.white, CLR_ROW_ALT]),
        ]
        if extra:
            s.extend(extra)
        return TableStyle(s)

    # -- header / footer callbacks --------------------------------------------
    def _header_footer(canvas, doc):
        canvas.saveState()
        width, height = A4

        # --- header separator ---
        canvas.setStrokeColor(CLR_BORDER)
        canvas.setLineWidth(0.5)
        canvas.line(20 * mm, height - 14 * mm, width - 20 * mm, height - 14 * mm)

        # --- footer ---
        footer_y = 10 * mm
        canvas.setStrokeColor(CLR_BORDER)
        canvas.line(20 * mm, footer_y + 4 * mm, width - 20 * mm, footer_y + 4 * mm)

        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(CLR_TEXT_SEC)
        canvas.drawString(
            20 * mm, footer_y,
            f"Documento gerado eletronicamente \u2022 Hash: {hash_doc}",
        )
        canvas.drawString(
            20 * mm, footer_y - 8,
            "Este relatório possui validade jurídica conforme MP 2.200-2/2001",
        )
        canvas.drawRightString(
            width - 20 * mm, footer_y - 8,
            f"Página {doc.page} \u2022 Gerado em {data_geracao}",
        )
        canvas.restoreState()

    # -- build elements -------------------------------------------------------
    elements: list = []

    # ---- CABEÇALHO ----------------------------------------------------------
    elements.append(Paragraph("RELATÓRIO FINANCEIRO", title_style))
    elements.append(Spacer(1, 2 * mm))

    elements.append(HRFlowable(
        width="100%",
        thickness=0.5,
        color=CLR_BORDER,
        spaceBefore=1 * mm,
        spaceAfter=1 * mm
    ))
    elements.append(Spacer(1, 2 * mm))

    elements.append(Paragraph(f"Documento Nº {hash_doc}", header_info_style))
    if req.customer_name:
        elements.append(Paragraph(f"Cliente: {req.customer_name}", header_info_style))
    if req.customer_cpf_cnpj:
        elements.append(Paragraph(
            f"CPF/CNPJ: {fmt_cpf_cnpj(req.customer_cpf_cnpj)}", header_info_style,
        ))
    elements.append(Paragraph(f"Período: {periodo}", header_info_style))
    elements.append(Paragraph(f"Data de Emissão: {data_geracao}", header_info_style))
    elements.append(Paragraph(
        f"Total de Operações: {len(req.transactions)}", header_info_style,
    ))

    # ---- SEÇÃO 1: RESUMO FINANCEIRO -----------------------------------------
    resumo_elementos = [
        Paragraph("1. RESUMO FINANCEIRO", section_style),
    ]

    resumo_data = [
        [
            Paragraph("Métrica", cell_bold),
            Paragraph("Valor (R$)", cell_bold_right),
        ],
        [
            Paragraph("Total de Entradas", cell_style),
            Paragraph(fmt_real(inflow_total), cell_right),
        ],
        [
            Paragraph("Total de Saídas", cell_style),
            Paragraph(fmt_real(outflow_total), cell_right),
        ],
        [
            Paragraph("Saldo do Período", cell_bold),
            Paragraph(fmt_real(saldo), cell_bold_right),
        ],
    ]

    resumo_table = Table(resumo_data, colWidths=[100 * mm, 70 * mm])
    resumo_extra = [
        ("ALIGN",      (1, 0), (1, -1), "RIGHT"),
        ("BACKGROUND", (0, 3), (-1, 3), CLR_SALDO_BG),
        ("FONTNAME",   (0, 3), (-1, 3), "Helvetica-Bold"),
    ]
    resumo_table.setStyle(_base_table_style(len(resumo_data), resumo_extra))
    resumo_elementos.append(resumo_table)
    elements.append(KeepTogether(resumo_elementos))

    # ---- SEÇÃO 2: DISTRIBUIÇÃO POR CATEGORIA --------------------------------
    cat_totals: dict[str, float] = {}
    for tx in req.transactions:
        if tx.transaction_type == "Outflow":
            cat_totals[tx.category] = cat_totals.get(tx.category, 0) + tx.total_amount

    if cat_totals:
        cat_elementos = [
            Paragraph("2. DISTRIBUIÇÃO POR CATEGORIA", section_style),
        ]

        grand_total = sum(cat_totals.values()) or 1.0
        cat_data = [[
            Paragraph("Categoria", cell_bold),
            Paragraph("Valor (R$)", cell_bold_right),
            Paragraph("% do Total", cell_bold_right),
        ]]
        for cat, total in sorted(cat_totals.items(), key=lambda x: -x[1]):
            pct = (total / grand_total) * 100
            cat_data.append([
                Paragraph(cat, cell_left),
                Paragraph(fmt_real(total), cell_right),
                Paragraph(f"{pct:.1f}%", cell_right),
            ])

        cat_table = Table(cat_data, colWidths=[60 * mm, 55 * mm, 40 * mm])
        cat_extra = [
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ]
        cat_table.setStyle(_base_table_style(len(cat_data), cat_extra))
        cat_elementos.append(cat_table)

        # Only use KeepTogether if the table is small enough
        if len(cat_data) <= 15:
            elements.append(KeepTogether(cat_elementos))
        else:
            elements.extend(cat_elementos)

    # ---- SEÇÃO 3: DETALHAMENTO DE TRANSAÇÕES --------------------------------
    elements.append(Paragraph("3. DETALHAMENTO DE TRANSAÇÕES", section_style))

    tx_header_row = [
        Paragraph("Data", cell_bold),
        Paragraph("Beneficiário/Pagador", cell_bold),
        Paragraph("Categoria", cell_bold),
        Paragraph("Natureza", cell_bold),
        Paragraph("Valor (R$)", cell_bold_right),
    ]
    tx_data = [tx_header_row]
    for tx in sorted_tx:
        dt = _parse_tx_date(tx.transaction_date)
        date_str = dt.strftime("%d/%m/%Y")
        merchant = tx.merchant_name[:30] if len(tx.merchant_name) > 30 else tx.merchant_name
        natureza = "Entrada" if tx.transaction_type == "Inflow" else "Saída"
        tx_data.append([
            Paragraph(date_str, cell_center),
            Paragraph(merchant, cell_left),
            Paragraph(tx.category, cell_left),
            Paragraph(natureza, cell_center),
            Paragraph(fmt_real(tx.total_amount), cell_right),
        ])

    col_widths = [25 * mm, 52 * mm, 30 * mm, 22 * mm, 35 * mm]
    tx_table = Table(tx_data, colWidths=col_widths, repeatRows=1)
    tx_extra = [
        ("ALIGN", (0, 0), (0, -1), "CENTER"),   # Data centralizada
        ("ALIGN", (3, 0), (3, -1), "CENTER"),    # Natureza centralizada
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),     # Valor à direita
    ]
    tx_table.setStyle(_base_table_style(len(tx_data), tx_extra))
    elements.append(tx_table)

    # ---- SEÇÃO 4: OBSERVAÇÕES E ASSINATURA ----------------------------------
    assinatura_data = [
        [''],  # Empty cell for the line
        [Paragraph("Assinatura Digital / Responsável", cell_center)]
    ]
    assinatura_table = Table(assinatura_data, colWidths=[100 * mm])
    assinatura_table.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (0, 0), 0.5, CLR_TEXT_PRI),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (0, 0), 5 * mm),
    ]))

    secao4_elementos = [
        Paragraph("4. OBSERVAÇÕES", section_style),
        Paragraph(
            "Este documento apresenta o resumo consolidado das movimentações "
            "financeiras do período indicado. Os valores estão expressos em "
            "Reais (R$) e seguem as normas contábeis brasileiras.",
            obs_style,
        ),
        Spacer(1, 15 * mm),
        assinatura_table
    ]
    elements.append(KeepTogether(secao4_elementos))

    # -- build PDF ------------------------------------------------------------
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=15 * mm, bottomMargin=15 * mm,
        leftMargin=20 * mm, rightMargin=20 * mm,
    )
    doc.build(elements, onFirstPage=_header_footer, onLaterPages=_header_footer)
    buf.seek(0)

    filename = f"relatorio_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
